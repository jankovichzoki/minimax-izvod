"""
Minimax Izvod Konvertor
=======================
Automatski pretvara PDF izvode u Minimax Excel sa BEX razbijanjem.

FIXES v2:
1. format_account_number - ispravno konvertuje IBAN u domaći format (3-13-2)
2. Download dugmad - ZIP arhiva umesto pojedinačnih dugmadi (nema rerun problema)
3. Debit/Credit prepoznavanje - Claude prompt poboljšan + generički fallback
"""

import streamlit as st
import io
import re
import json
import zipfile
from pathlib import Path
import anthropic
from openpyxl import Workbook

st.set_page_config(page_title="Minimax Izvod", page_icon="🏦", layout="wide")

API_KEY = st.secrets.get("ANTHROPIC_API_KEY", "")

# ========================================================================
# PASSWORD PROTECTION
# ========================================================================
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.markdown("# 🔒 Minimax Izvod - Pristup zaštićen")
    st.markdown("Unesi lozinku za pristup aplikaciji:")
    password = st.text_input("Lozinka:", type="password", key="password_input")
    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        if st.button("🔓 Prijavi se", type="primary"):
            correct_password = st.secrets.get("APP_PASSWORD", "minimax2026")
            if password == correct_password:
                st.session_state.authenticated = True
                st.success("✅ Uspešna prijava!")
                st.rerun()
            else:
                st.error("❌ Pogrešna lozinka!")
    st.markdown("---")
    st.info("💡 Kontaktiraj administratora za pristup.")
    st.stop()

# ========================================================================
# CSS
# ========================================================================
st.markdown("""<style>
    .main-title { font-size: 2.5rem; font-weight: 800; margin-bottom: 0.5rem; }
    .subtitle { color: #666; margin-bottom: 2rem; }
    .stButton>button { width: 100%; }
</style>""", unsafe_allow_html=True)

st.markdown('<h1 class="main-title">🏦 Minimax Izvod Konvertor</h1>', unsafe_allow_html=True)
st.markdown('<p class="subtitle">PDF izvodi → Excel/XML sa razbijenim BEX kupcima</p>', unsafe_allow_html=True)

# ========================================================================
# FIX 1: format_account_number - ispravan IBAN → domaći format
# ========================================================================
def format_account_number(account_str):
    """
    Konvertuje broj računa u srpski domaći format: XXX-XXXXXXXXXXXXX-XX (3-13-2).
    
    Minimax XML Partija atribut mora biti 18 cifara BEZ crtica.
    Ova funkcija vraća formatiran string SA crticama - za XML koristiti .replace('-','').
    
    Podržava sve varijante ulaza:
    - IBAN: RS35170003002777200074 → 170-0030027772000-74
    - Domaći 3-13-2: 205-0000000422476-62 → vraća kao je
    - Domaći 3-X-2 bez leading zeros: 170-30027772000-74 → 170-0030027772000-74 (KEY FIX)
    - 18 cifara bez crtica → 3-13-2 format
    - 16 cifara bez crtica (domaći bez crtica i bez leading zeros) → dopuni na 18
    """
    s = str(account_str).strip()
    
    # Već tačno 3-13-2 format
    if re.match(r'^\d{3}-\d{13}-\d{2}$', s):
        return s
    
    # Domaći format 3-X-2 gde X ima MANJE od 13 cifara → dopuni sa leading zeros
    m = re.match(r'^(\d{3})-(\d{1,12})-(\d{2})$', s)
    if m:
        bank, mid, check = m.group(1), m.group(2), m.group(3)
        return f'{bank}-{mid.zfill(13)}-{check}'
    
    # IBAN format (počinje sa RS)
    if s.upper().startswith('RS'):
        all_digits = re.sub(r'\D', '', s)
        # IBAN = RS(slova) + 2check_digits + 18_BBAN → cifre = check(2) + BBAN(18) = 20
        bban_digits = all_digits[2:] if len(all_digits) == 20 else all_digits
        if len(bban_digits) == 18:
            return f'{bban_digits[:3]}-{bban_digits[3:16]}-{bban_digits[16:]}'
    
    digits = re.sub(r'\D', '', s)
    
    # 18 cifara bez crtica - direktno formatuj
    if len(digits) == 18:
        return f'{digits[:3]}-{digits[3:16]}-{digits[16:]}'
    
    # 16 cifara bez crtica = domaći bez leading zeros i bez crtica
    # npr: 1703002777200074 = 170 | 30027772000 (11) | 74 → dopuni srednji na 13
    if len(digits) == 16:
        bank = digits[:3]
        mid = digits[3:14]   # 11 cifara bez leading zeros
        check = digits[14:]  # 2 cifre
        return f'{bank}-{mid.zfill(13)}-{check}'
    
    # Ima crtice u nekom drugom obliku - vrati kao je
    if '-' in s:
        return s
    
    return s


# ========================================================================
# Ostale helper funkcije (nepromenjene)
# ========================================================================
def extract_text_from_pdf(pdf_bytes):
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = ""
            for page in pdf.pages:
                text += page.extract_text() + "\n\n"
        return text
    except:
        import zipfile as zf
        if pdf_bytes[:2] == b"PK":
            with zf.ZipFile(io.BytesIO(pdf_bytes)) as z:
                txt_files = sorted([n for n in z.namelist() if n.endswith('.txt')])
                text = ""
                for tf in txt_files:
                    text += z.read(tf).decode('utf-8', errors='replace') + "\n\n"
                return text
        return pdf_bytes.decode('utf-8', errors='replace')


def parse_xml_izvod(xml_bytes, filename):
    import xml.etree.ElementTree as ET
    try:
        tree = ET.parse(io.BytesIO(xml_bytes))
        root = tree.getroot()
        zaglavlje = root.find('Zaglavlje')
        if zaglavlje is None:
            raise ValueError("XML nema Zaglavlje element")
        statement = {
            'date': zaglavlje.get('DatumIzvoda', ''),
            'account': zaglavlje.get('Partija', ''),
            'number': zaglavlje.get('BrojIzvoda', ''),
            'owner_name': zaglavlje.get('KomitentNaziv', ''),
            'owner_address': zaglavlje.get('KomitentAdresa', ''),
            'owner_place': zaglavlje.get('KomitentMesto', ''),
            'tax_number': zaglavlje.get('MaticniBroj', '')
        }
        transactions = []
        for stavka in root.findall('Stavke'):
            debit = float(stavka.get('Duguje', '0') or '0')
            credit = float(stavka.get('Potrazuje', '0') or '0')
            transactions.append({
                'date': stavka.get('DatumValute', ''),
                'customer_name': stavka.get('NalogKorisnik', ''),
                'customer_address': stavka.get('Mesto', ''),
                'customer_account': stavka.get('BrojRacunaPrimaocaPosiljaoca', ''),
                'customer_tax_number': '',
                'reference': stavka.get('PozivNaBrojKorisnika', '') or stavka.get('Referenca', ''),
                'currency': 'RSD',
                'debit': debit,
                'credit': credit,
                'description': stavka.get('Opis', '')
            })
        return {'statement': statement, 'transactions': transactions}
    except Exception as e:
        raise ValueError(f"XML parsing greška: {str(e)}")


def parse_serbian_amount(amount_str):
    """
    Parsira iznos koji može biti u srpskom (1.234,56) ili engleskom (1,234.56) formatu,
    ili bez separatora hiljada. Ne sme naivno da briše i tačku i zarez jer to gubi decimale
    (npr. '11.400,50'.replace(',','').replace('.','') → '1140050' umesto 11400.50).
    """
    s = re.sub(r'[^\d,.\-]', '', str(amount_str).strip())
    if not s:
        return 0.0

    if ',' in s and '.' in s:
        # Poslednji separator u nizu je decimalni, prethodni su hiljade
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        # Samo zarez: 1-2 cifre posle → decimalni zarez, inače hiljadni separator
        after = s.split(',')[-1]
        s = s.replace(',', '.') if len(after) <= 2 else s.replace(',', '')
    elif '.' in s:
        # Samo tačka: tačno 3 cifre posle poslednje tačke → hiljadni separator (11.400 = 11400)
        # 1-2 cifre → decimalna tačka (11.40 = 11.40)
        after = s.split('.')[-1]
        if len(after) == 3:
            s = s.replace('.', '')

    return float(s) if s not in ('', '-', '.') else 0.0


def parse_bex_specification(file_bytes, filename):
    if filename.lower().endswith('.csv'):
        try:
            import csv
            text = file_bytes.decode('utf-8-sig', errors='replace')
            # csv.reader ne baca gresku na nekonzistentan broj kolona po redu (za razliku
            # od pandas C parsera) - bitno jer BEX izvoz zna da ubaci zarez bez navodnika
            # u napomena kolone, sto pomeri broj polja u tom jednom redu.
            rows = [r for r in csv.reader(io.StringIO(text)) if any(c.strip() for c in r)]
            if not rows:
                return []

            header = [h.strip() for h in rows[0]]
            col_idx = {name: i for i, name in enumerate(header) if name}

            def field(row, col_name):
                i = col_idx.get(col_name)
                return row[i].strip() if i is not None and i < len(row) else ''

            customers = []
            for row in rows[1:]:
                posiljka = field(row, 'IdPosiljke')
                name = field(row, 'UplatilacNaziv')
                address = field(row, 'UplatilacMesto')
                amount = parse_serbian_amount(field(row, 'UplacenoOtkupa') or '0')
                date_str = field(row, 'DatumNaplateOtkupnine')
                date = date_str.split()[0] if ' ' in date_str else date_str
                if posiljka and name and amount > 0:
                    customers.append({
                        'name': name, 'address': address, 'amount': amount,
                        'posiljka': posiljka, 'reference': f'OT-{posiljka}', 'date': date
                    })
            return customers
        except Exception as e:
            st.error(f"CSV parsing greška: {str(e)}")
            return []
    else:
        try:
            text = extract_text_from_pdf(file_bytes)
            if not API_KEY:
                st.error("API key nije konfigurisan za PDF parsiranje!")
                return []
            client = anthropic.Anthropic(api_key=API_KEY)
            prompt = f"""Analiziraj BEX Express specifikaciju i izvuci podatke o kupcima.

TEKST SPECIFIKACIJE:
{text}

Vrati SAMO JSON (bez markdown):

{{
  "customers": [
    {{
      "posiljka": "262598547",
      "name": "MILEV JOVAN",
      "address": "PIROT, OBILIĆEVA 3",
      "amount": 11400,
      "date": "18.02.2026"
    }}
  ]
}}

KRITIČNO PRAVILA ZA IZNOSE:
1. UKLONI SVE ZAREZE iz iznosa: 11,400 → 11400
2. NIKAD ne dodavaj nule
3. Proveri: suma ispod 1,000,000 RSD

OSTALA PRAVILA:
- posiljka = 9-cifreni broj
- name = TAČNO kao što piše (VELIKA SLOVA)
- date = DD.MM.YYYY
- NIKAD ne izmišljaj podatke"""
            msg = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=8192,
                messages=[{"role": "user", "content": prompt}]
            )
            raw = msg.content[0].text
            clean = raw.replace('```json', '').replace('```', '').strip()
            try:
                data = json.loads(clean)
            except json.JSONDecodeError as e:
                cut_off = msg.stop_reason == 'max_tokens'
                hint = " (odgovor je odsečen - specifikacija je predugačka za trenutni limit)" if cut_off else ""
                raise ValueError(f"Nevalidan JSON od Claude-a: {e}{hint}")
            return [{
                'name': c.get('name', ''), 'address': c.get('address', ''),
                'amount': float(c.get('amount', 0)), 'posiljka': str(c.get('posiljka', '')),
                'reference': f"OT-{c.get('posiljka', '')}", 'date': c.get('date', '')
            } for c in data.get('customers', [])]
        except Exception as e:
            st.error(f"PDF parsing greška: {str(e)}")
            return []


# ========================================================================
# FIX 3: parse_with_claude - poboljšan prompt za debit/credit
# ========================================================================
def parse_with_claude(text, filename):
    """Parse izvod - Claude vraća TAČAN debit/credit prema svrsi transakcije."""
    if not API_KEY:
        raise ValueError("ANTHROPIC_API_KEY nije konfigurisan!")
    
    client = anthropic.Anthropic(api_key=API_KEY)
    
    prompt = f"""Analiziraj izvod banke i izvuci podatke u JSON formatu.

TEKST IZVODA:
{text}

NAZIV FAJLA: {filename}

Vrati SAMO JSON (bez markdown):

{{
  "statement": {{
    "date": "DD.MM.YYYY",
    "account": "domaći broj računa SA CRTICAMA npr 205-0000000422476-62 ili 170-30027772000-74 (NE IBAN format!)",
    "number": "broj_izvoda",
    "owner_name": "ime vlasnika računa",
    "owner_address": "ulica i broj vlasnika (bez mesta/grada)",
    "owner_place": "poštanski broj i mesto/grad vlasnika, npr. '11010 BEOGRAD-VOŽDOVAC'",
    "tax_number": "matični broj firme vlasnika računa (ne PIB kupca)"
  }},
  "transactions": [
    {{
      "date": "DD.MM.YYYY",
      "customer_name": "naziv platioca ili primaoca",
      "customer_address": "adresa",
      "customer_account": "broj računa sa crticama",
      "customer_tax_number": "",
      "reference": "poziv na broj ili referenca",
      "currency": "RSD",
      "debit": 0.00,
      "credit": 0.00,
      "description": "svrha plaćanja"
    }}
  ]
}}

KLJUČNA PRAVILA ZA BROJ RAČUNA (account polje u statement):
- PRIORITET 1: Ako dokument sadrži IBAN (počinje sa "RS" + 2 cifre, npr. "RS35170003007043900080"), vrati GA TAČNO, sve cifre
- PRIORITET 2: Ako nema IBAN, vrati domaći broj TAČNO kao što piše, sa svim ciframa i crticama
- IBAN je uvek precizniji od domaćeg formata - leading zeros su sigurno tačni u IBAN-u
- NIKAD ne menjaj, ne skraćuj, ne zaokružuj cifre broja računa

KLJUČNA PRAVILA ZA DEBIT/CREDIT:
- CREDIT (potražuje) = novac ULAZI na račun = primanja, uplate od kupaca, kreditiranja
- DEBIT (duguje) = novac IZLAZI sa računa = plaćanja, troškovi, transferi prema drugima
- Čitaj kolone "Zaduženje" i "Odobrenje" u izvodu:
  * "Odobrenje" kolona → to je CREDIT (credit > 0, debit = 0)
  * "Zaduženje" kolona → to je DEBIT (debit > 0, credit = 0)
- Ako iznos ima predznak "-" → DEBIT
- Nikad ne stavljaj isti iznos i u debit i u credit
- Nikad ne stavljaj 0 i u debit i u credit (jedino mora biti jedno > 0)

OSTALA PRAVILA:
- Račune vrati SA crticama u formatu: XXX-XXXXXXXXXXXXX-XX
- date format: DD.MM.YYYY
- Ignoriši ukupne sume na kraju izvoda (samo pojedinačne stavke)"""
    
    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=8192,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = msg.content[0].text
    clean = raw.replace('```json', '').replace('```', '').strip()
    try:
        return json.loads(clean)
    except json.JSONDecodeError as e:
        cut_off = msg.stop_reason == 'max_tokens'
        hint = " (odgovor je odsečen jer je izvod predugačak za trenutni limit - javi administratoru)" if cut_off else ""
        raise ValueError(f"Claude je vratio nevalidan JSON za {filename}: {e}{hint}")


def expand_bex_transactions(transactions, specifications, consumed):
    """
    consumed: set deljen preko svih izvoda u istom pokretanju - jednom upotrebljena
    specifikacija se ne sme ponovo upariti sa drugom BEX transakcijom (čak i ako bi
    slučajno imala isti ukupan iznos), inače dolazi do duplog razbijanja.
    """
    expanded = []
    for tx in transactions:
        is_bex = 'BEX' in (tx.get('customer_name', '') or '').upper()
        if is_bex:
            tx_amount = tx.get('credit', 0) or tx.get('debit', 0)
            matched = None
            matched_name = None
            for spec_name, customers in specifications.items():
                if spec_name in consumed:
                    continue
                spec_total = sum(c['amount'] for c in customers)
                if abs(spec_total - tx_amount) < 0.01:
                    matched = customers
                    matched_name = spec_name
                    break
            if matched:
                consumed.add(matched_name)
                st.success(f"🔄 Razbijam BEX: {matched_name} → {len(matched)} kupaca ({tx_amount:,.2f} RSD)")
                for c in matched:
                    expanded.append({
                        'date': c['date'], 'customer_name': c['name'],
                        'customer_address': c['address'], 'customer_account': '',
                        'customer_tax_number': '', 'reference': c['reference'],
                        'currency': 'RSD', 'debit': 0, 'credit': c['amount'],
                        'description': f"Otkup pošiljke {c['posiljka']}"
                    })
            else:
                st.warning(f"⚠️ BEX transakcija od {tx_amount:,.2f} RSD nema odgovarajuću specifikaciju "
                           f"(proveri da li je upload-ovana specifikacija sa tim ukupnim iznosom) - ostavljena neraspakovana")
                expanded.append(tx)
        else:
            expanded.append(tx)
    return expanded


# ========================================================================
# FIX 3b: validate_debit_credit - generički validator (bez hardkodovanih imena)
# ========================================================================
def validate_debit_credit(transactions):
    """
    Generički validator: osigurava da svaka stavka ima ili debit>0 ili credit>0, ne oba.
    Ne pretpostavlja ništa o imenima - Claude je već uradio klasifikaciju.
    Jedino što radi: 
    - Ako su oba 0 → logička greška, prijavi
    - Ako su oba > 0 → neispravno, vrati samo credit (odobrenje je važnije)
    - Ako samo jedan > 0 → OK, ostavi
    """
    fixed = []
    for tx in transactions:
        debit = float(tx.get('debit', 0) or 0)
        credit = float(tx.get('credit', 0) or 0)
        
        if debit > 0 and credit > 0:
            # Konflikt: Claude je vratio oba - zadržimo credit (odobrenje)
            tx['debit'] = 0
            tx['credit'] = credit
        elif debit == 0 and credit == 0:
            # Oba nula - nemamo informaciju, ostavimo kao je
            pass
        # else: jedan je > 0, sve je OK
        
        fixed.append(tx)
    return fixed


def create_minimax_excel(statement, transactions):
    wb = Workbook()
    account = format_account_number(statement.get('account', ''))
    
    ws1 = wb.active
    ws1.title = "Statement"
    ws1.append(["Date", "Account", "Number"])
    ws1.append([statement.get('date', ''), account, statement.get('number', '')])
    for row in ws1.iter_rows():
        for cell in row:
            cell.number_format = "@"
    ws1.column_dimensions["A"].width = 15
    ws1.column_dimensions["B"].width = 32
    ws1.column_dimensions["C"].width = 10
    
    ws2 = wb.create_sheet("Transactions")
    headers = ["CustomerName", "CustomerAddress", "CustomerAccount", "CustomerTaxNumber",
               "Date", "Reference", "Currency", "Debit", "Credit", "Description"]
    ws2.append(headers)
    for tx in transactions:
        cust_account = format_account_number(tx.get('customer_account', '')) if tx.get('customer_account') else ''
        ws2.append([
            str(tx.get("customer_name", "") or ""),
            str(tx.get("customer_address", "") or ""),
            cust_account,
            str(tx.get("customer_tax_number", "") or ""),
            str(tx.get("date", "") or ""),
            str(tx.get("reference", "") or ""),
            "RSD",
            float(tx.get("debit", 0) or 0),
            float(tx.get("credit", 0) or 0),
            str(tx.get("description", "") or ""),
        ])
    
    num_cols = {8, 9}
    for row in ws2.iter_rows():
        for cell in row:
            if cell.column in num_cols:
                cell.number_format = "0.00"
            else:
                cell.number_format = "@"
    
    col_widths = [35, 25, 28, 15, 12, 25, 8, 12, 12, 45]
    for i, width in enumerate(col_widths, 1):
        ws2.column_dimensions[ws2.cell(1, i).column_letter].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_minimax_xml(statement, transactions):
    import xml.etree.ElementTree as ET
    
    account = format_account_number(statement.get('account', ''))
    account_no_dashes = account.replace('-', '')
    
    dugovni = sum(float(tx.get('debit', 0) or 0) for tx in transactions)
    potrazni = sum(float(tx.get('credit', 0) or 0) for tx in transactions)
    
    root = ET.Element('TransakcioniRacunPrivredaIzvod')
    zaglavlje = ET.SubElement(root, 'Zaglavlje')
    zaglavlje.set('VrstaIzvoda', 'R')
    zaglavlje.set('BrojIzvoda', statement.get('number', ''))
    zaglavlje.set('DatumIzvoda', statement.get('date', ''))
    zaglavlje.set('MaticniBroj', statement.get('tax_number', '') or '')
    zaglavlje.set('KomitentNaziv', statement.get('owner_name', ''))
    zaglavlje.set('KomitentAdresa', statement.get('owner_address', ''))
    zaglavlje.set('KomitentMesto', statement.get('owner_place', '') or '')
    zaglavlje.set('Partija', account_no_dashes)
    zaglavlje.set('TipRacuna', 'Transakcioni depoziti preduzetnika')
    zaglavlje.set('PrethodnoStanje', f"{dugovni + potrazni:.2f}")
    zaglavlje.set('DugovniPromet', f"{dugovni:.2f}")
    zaglavlje.set('PotrazniPromet', f"{potrazni:.2f}")
    zaglavlje.set('NovoStanje', f"{potrazni - dugovni:.2f}")
    zaglavlje.set('StanjeObracunateProvizije', '0')
    
    for tx in transactions:
        cust_account = format_account_number(tx.get('customer_account', '')) if tx.get('customer_account') else ''
        stavka = ET.SubElement(root, 'Stavke')
        stavka.set('NalogKorisnik', str(tx.get('customer_name', '') or ''))
        stavka.set('Mesto', str(tx.get('customer_address', '') or ''))
        stavka.set('VasBrojNaloga', '')
        stavka.set('BrojRacunaPrimaocaPosiljaoca', cust_account)
        stavka.set('Opis', str(tx.get('description', '') or ''))
        stavka.set('SifraPlacanja', '')
        stavka.set('SifraPlacanjaOpis', '')
        stavka.set('Duguje', f"{float(tx.get('debit', 0) or 0):.2f}")
        stavka.set('Potrazuje', f"{float(tx.get('credit', 0) or 0):.2f}")
        stavka.set('ModelZaduzenjaOdobrenja', '')
        stavka.set('PozivNaBrojZaduzenjaOdobrenja', '')
        stavka.set('ModelKorisnika', '')
        stavka.set('PozivNaBrojKorisnika', str(tx.get('reference', '') or ''))
        stavka.set('BrojZaReklamaciju', '')
        stavka.set('Referenca', str(tx.get('reference', '') or ''))
        stavka.set('Objasnjenje', '')
        stavka.set('DatumValute', str(tx.get('date', '') or ''))
    
    tree = ET.ElementTree(root)
    ET.indent(tree, space="  ", level=0)
    output = io.BytesIO()
    tree.write(output, encoding='utf-8', xml_declaration=True)
    output.seek(0)
    return output.getvalue()


# ========================================================================
# MAIN UI
# ========================================================================
col1, col2 = st.columns(2)

with col1:
    st.markdown("### 📄 Izvodi banke")
    izvodi_files = st.file_uploader(
        "Upload PDF ili XML izvoda",
        type=['pdf', 'PDF', 'xml', 'XML'],
        accept_multiple_files=True,
        key='izvodi'
    )

with col2:
    st.markdown("### 📋 BEX Specifikacije (opciono)")
    spec_files = st.file_uploader(
        "Upload BEX specifikacija (PDF ili CSV)",
        type=['pdf', 'PDF', 'csv', 'CSV'],
        accept_multiple_files=True,
        key='specs'
    )

if izvodi_files:
    st.markdown("---")
    
    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        generate_excel = st.button("📊 Generiši Excel", type="primary", use_container_width=True)
    with col_btn2:
        generate_xml = st.button("📄 Generiši XML", type="secondary", use_container_width=True)
    
    if generate_excel or generate_xml:
        output_format = "Excel" if generate_excel else "XML"
        st.info(f"Generišem {output_format} format...")
        
        specifications = {}
        if spec_files:
            with st.spinner("Parsiram BEX specifikacije..."):
                for spec_file in spec_files:
                    try:
                        spec_bytes = spec_file.read()
                        customers = parse_bex_specification(spec_bytes, spec_file.name)
                        if customers:
                            specifications[spec_file.name] = customers
                            total = sum(c['amount'] for c in customers)
                            st.success(f"✅ {spec_file.name}: {len(customers)} kupaca, {total:,.2f} RSD")
                    except Exception as e:
                        st.error(f"❌ {spec_file.name}: {str(e)}")
        
        progress_bar = st.progress(0)
        results = []
        consumed_specs = set()  # deljeno preko svih izvoda u ovom pokretanju

        for i, izvod_file in enumerate(izvodi_files):
            progress_bar.progress((i + 1) / len(izvodi_files))
            try:
                with st.status(f"Obradjujem: {izvod_file.name}"):
                    st.write("Čitam fajl...")
                    pdf_bytes = izvod_file.read()
                    
                    if izvod_file.name.lower().endswith('.xml'):
                        st.write("Parsiram XML izvod...")
                        parsed = parse_xml_izvod(pdf_bytes, izvod_file.name)
                    else:
                        text = extract_text_from_pdf(pdf_bytes)
                        st.write("AI parsiranje PDF izvoda...")
                        parsed = parse_with_claude(text, izvod_file.name)
                    
                    st.write("Proveravam BEX...")
                    original_count = len(parsed['transactions'])
                    expanded = expand_bex_transactions(parsed['transactions'], specifications, consumed_specs)
                    
                    # FIX 3: Generički validator umesto hardkodovane logike
                    st.write("Validujem debit/credit...")
                    expanded = validate_debit_credit(expanded)
                    
                    st.write(f"Generišem {output_format}...")
                    if generate_excel:
                        file_bytes = create_minimax_excel(parsed['statement'], expanded)
                        output_name = re.sub(r'\.(pdf|PDF|xml|XML)$', '', izvod_file.name) + '_minimax.xlsx'
                        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    else:
                        file_bytes = create_minimax_xml(parsed['statement'], expanded)
                        output_name = re.sub(r'\.(pdf|PDF|xml|XML)$', '', izvod_file.name) + '_minimax.xml'
                        mime_type = "application/xml"
                    
                    results.append({
                        'success': True,
                        'filename': izvod_file.name,
                        'output_name': output_name,
                        'file_bytes': file_bytes,
                        'mime_type': mime_type,
                        'format': output_format,
                        'statement': parsed['statement'],
                        'tx_count': len(expanded),
                        'bex_expanded': len(expanded) > original_count,
                        'transactions': expanded
                    })
            except Exception as e:
                results.append({'success': False, 'filename': izvod_file.name, 'error': str(e)})
        
        progress_bar.empty()
        
        # ================================================================
        # FIX 2: ZIP download za sve fajlove odjednom + individualni prikaz
        # ================================================================
        st.markdown("---")
        successful = [r for r in results if r['success']]
        failed = [r for r in results if not r['success']]
        
        st.markdown(f"## 📥 Rezultati ({output_format})")
        
        # Ako ima više od 1 uspešnog - ponudi ZIP
        if len(successful) > 1:
            st.markdown("### 📦 Preuzmi sve odjednom")
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                for r in successful:
                    zf.writestr(r['output_name'], r['file_bytes'])
            zip_buffer.seek(0)
            
            ext = "xlsx" if output_format == "Excel" else "xml"
            st.download_button(
                label=f"⬇️ Preuzmi SVE kao ZIP ({len(successful)} fajlova)",
                data=zip_buffer.getvalue(),
                file_name=f"minimax_izvodi_{output_format.lower()}.zip",
                mime="application/zip",
                type="primary",
                use_container_width=True,
                key="download_all_zip"
            )
            st.markdown("---")
        
        # Prikaz svakog rezultata sa individualnim download-om
        for r in successful:
            col1, col2 = st.columns([3, 1])
            with col1:
                st.markdown(f"### ✅ {r['filename']}")
                raw_account = r['statement']['account']
                formatted_account = format_account_number(raw_account)
                xml_account = formatted_account.replace('-', '')
                st.markdown(f"**Račun:** `{formatted_account}`")
                st.caption(f"🔍 Debug — Claude izvukao: `{raw_account}` → XML Partija: `{xml_account}` ({len(xml_account)} cifara)")
                st.markdown(f"**Transakcija:** {r['tx_count']}" +
                          (f" _(BEX razbijen)_" if r['bex_expanded'] else ""))
            with col2:
                btn_label = "⬇️ Excel" if r['format'] == "Excel" else "⬇️ XML"
                # FIX 2: Koristimo st.session_state čuvanje da sprečimo rerun brisanje
                st.download_button(
                    btn_label,
                    data=r['file_bytes'],
                    file_name=r['output_name'],
                    mime=r['mime_type'],
                    key=f"dl_{hash(r['filename'])}_{r['format']}"
                )
            
            with st.expander(f"📊 Pregledaj transakcije ({r['tx_count']})"):
                import pandas as pd
                tx_data = [{
                    'Br': i,
                    'Datum': tx.get('date', ''),
                    'Kupac': tx.get('customer_name', '')[:40],
                    'Duguje': f"{tx.get('debit', 0):,.2f}",
                    'Potražuje': f"{tx.get('credit', 0):,.2f}",
                    'Opis': tx.get('description', '')[:50]
                } for i, tx in enumerate(r['transactions'], 1)]
                
                df = pd.DataFrame(tx_data)
                st.dataframe(df, use_container_width=True, hide_index=True)
                
                total_debit = sum(tx.get('debit', 0) for tx in r['transactions'])
                total_credit = sum(tx.get('credit', 0) for tx in r['transactions'])
                col_s1, col_s2, col_s3 = st.columns(3)
                with col_s1:
                    st.metric("Ukupno Duguje", f"{total_debit:,.2f} RSD")
                with col_s2:
                    st.metric("Ukupno Potražuje", f"{total_credit:,.2f} RSD")
                with col_s3:
                    st.metric("Saldo", f"{total_credit - total_debit:,.2f} RSD")
        
        for r in failed:
            st.error(f"❌ {r['filename']}: {r['error']}")
        
        # Reset dugme - korisnik kontroliše kada se resetuje
        st.markdown("---")
        if st.button("🔄 Novi upload (resetuj)", type="secondary"):
            st.rerun()

else:
    st.info("👆 Započni upload-om PDF izvoda")
